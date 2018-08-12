import random
from itertools import combinations

import ngram
import openpyxl
from pyhanlp import HanLP

"""
@Author: xiezizhe
@Date: 2018/8/10 下午3:24
"""


def open_excel(file_name, sheet_idx=None):
    """
    \
    Args:
        file_name:
        sheet_idx:

    Returns:

    """
    wb = openpyxl.load_workbook(file_name)
    if sheet_idx is None:
        sheet_idx = 0
    return wb[wb.sheetnames[0]]


def replace_str(str):
    try:
        return str.replace("\r", "").replace("?", "").replace("？", "").strip()
    except Exception:
        return str


def generate_file(file_name, output_file_name):
    sheet_input = open_excel(file_name=file_name)
    end_pos = 10000
    faq_data = dict()
    standard_question_idx = 0
    standard_question_to_id = dict()
    for i in range(2, end_pos):
        if sheet_input.cell(i, 1).value is None and sheet_input.cell(i, 2).value is None:
            break

        standard_question = replace_str(sheet_input.cell(i, 4).value)
        if standard_question is None or sheet_input.cell(i, 5).value is None:
            continue
        if standard_question not in standard_question_to_id:
            standard_question_to_id[standard_question] = standard_question_idx
            standard_question_idx += 1
        if standard_question not in faq_data:
            faq_data[standard_question] = []
        sim_questions = sheet_input.cell(i, 5).value.split('\n')
        sim_questions.append(standard_question)
        faq_data[standard_question].extend(
            list(set(
                map(replace_str, filter(lambda x: x != "", sim_questions)))))

    print(len(faq_data))
    faq_keys = dict()
    for i, (k, vs) in enumerate(faq_data.items()):
        for v in vs:
            faq_keys[v] = standard_question_to_id[k]
    k = 30
    total_groups = len(faq_data)
    training_data = []
    for i, key in enumerate(list(faq_data.keys())):
        if i % 1 == 0:
            print("{}/{}".format(i, total_groups))
        label = standard_question_to_id[key]
        questions = faq_data[key]
        corpus_list = []
        for items in filter(lambda x: x[0] != key, faq_data.items()):
            corpus_list.extend(items[1])
        if questions[0] in corpus_list:
            print("impossible")
        G = ngram.NGram(corpus_list, N=2)
        cache = dict()
        for positive_question, anchor_question in combinations(questions, 2):
            if anchor_question in cache:
                negative_questions = cache[anchor_question]
            else:
                negative_questions = G.search(anchor_question, threshold=0.1)
                cache[anchor_question] = negative_questions
            ans_len = min(len(negative_questions), 30)
            for i in range(ans_len):
                negative_question = negative_questions[i][0]
                training_data.append(
                    "{} {} {}\t{}\t{}\t{}".format(label, label, faq_keys[negative_question], positive_question,
                                                  anchor_question, negative_question))
    with open(output_file_name, "w", encoding="utf-8") as fw:
        for value in training_data:
            fw.write(value)
            fw.write("\r\n")


def shuffle_file(output_file_name):
    training_data = []
    with open(output_file_name, "r", encoding="utf-8") as fr:
        for line in fr.readlines():
            if line is not None and line.strip() != "":
                training_data.append(line.strip())
    print(len(training_data))
    random.shuffle(training_data)
    print("start to rewrite file")
    with open(output_file_name, 'w', encoding="utf-8") as fw:
        for value in training_data:
            fw.write(value)
            fw.write("\n")
    print("Done.")


def get_embedding(word2idx, sentence, cache):
    embedding = []
    if sentence in cache:
        return cache[sentence]
    for term in HanLP.segment(sentence):
        if term.word in word2idx:
            embedding.append(word2idx[term.word])
    if len(embedding) > 25:
        embedding = embedding[:24]
    while len(embedding) < 25:
        embedding.append(-1)
    value = " ".join(map(str, embedding))
    cache[sentence] = value
    return value


def tokenize_embedding(output_file_name):
    word2idx = dict()
    total_num = 0
    for index, line in enumerate(open('../data/model.vec', 'r', encoding="utf-8")):
        word_embedding = line.split(' ')
        total_num += 1
        if len(word_embedding) >= 256:
            word2idx[word_embedding[0]] = index - 1
    print("{} lines in total".format(total_num))
    cache = dict()
    with open(output_file_name, "r", encoding="utf-8") as fr:
        with open(output_file_name.replace(".txt", "_tokenize.txt"), 'w', encoding="utf-8") as fw:
            for index, line in enumerate(fr.readlines()):
                if index % 100000 == 0:
                    fw.flush()
                    print("read {} lines".format(index))
                try:
                    array = line.split('\t')
                    fw.write("{}\t{}\t{}\t{}".format(array[0], get_embedding(word2idx, array[1], cache),
                                                     get_embedding(word2idx, array[2], cache),
                                                     get_embedding(word2idx, array[3], cache)))
                    fw.write('\n')
                except Exception:
                    print(line)

    print("Exit tokenize_embedding")


if __name__ == '__main__':
    excel_file_name = '../data/faq_train_90_.xlsx'
    output_file_name = '../data/faq_training.txt'
    tokenize_embedding(output_file_name=output_file_name)
    # if Path(output_file_name).is_file():
    #     shuffle_file(output_file_name=output_file_name)
    # else:
    #     generate_file(file_name=excel_file_name, output_file_name=output_file_name)
    print("Hello, world")
